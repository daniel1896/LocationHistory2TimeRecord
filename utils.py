import math
import openpyxl


def create_excel_file(path, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    # add column names
    ws['A1'] = 'Date'
    ws['B1'] = 'Start Time'
    ws['C1'] = 'End Time'
    ws['D1'] = 'Duration'
    # save the file
    wb.save(path)
    return wb, ws


def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371 * 1000  # radius of the Earth in meters

    # Convert latitude and longitude to radians
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)

    # Haversine formula
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    distance = R * c

    return distance


def is_within_radius(target_lat, target_lon, lat, lon, radius):
    distance = calculate_distance(target_lat, target_lon, lat, lon)
    return distance <= radius


if __name__ == '__main__':
    # ws = create_excel_file('test.xlsx', 'test')

    # Example usage
    latitude, longitude = 53.85723631711899, 10.672369351617364
    target_latitude, target_longitude = 53.85798340637453, 10.673058504614588
    target_radius = 100  # 100 meters

    if is_within_radius(target_latitude, target_longitude, latitude, longitude, target_radius):
        print("The coordinate is within the radius.")
    else:
        print("The coordinate is outside the radius.")
